"""
running_order.py
Manages the Running Order — the row-based instruction list that drives report assembly —
as a list of row dicts, with .xlsx export/import for human editing.
"""

import os
import csv
import json

from packages.constants_temp.constants_temp import coerce_row

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

COLUMNS = [
    "row_id",
    "enabled",
    "scope",
    "function",
    "slide_index",
    "placeholder",
    "chart_type_ref",
    "cache_file",
    "populations",
    "image_path",
    "excel_path",
    "export_range",
    "driver_range",
    "left_emu",
    "top_emu",
    "width_emu",
    "height_emu",
    "notes",
]

ALL_FUNCTIONS = [
    "create_ppt",
    "set_default_populations",
    "update_text",
    "insert_chart",
    "insert_picture",
    "insert_from_excel",
    "open_excel",
    "close_excel",
    "empty_placeholder",
    "save_ppt",
    "save_pdf",
]

SCOPE_VALUES = ["normal", "batch_open", "batch_close"]

STRUCTURAL_FUNCTIONS = {"create_ppt", "set_default_populations", "update_text", "save_ppt", "save_pdf"}
CONTENT_FUNCTIONS    = {"insert_chart", "insert_picture", "insert_from_excel", "empty_placeholder"}
BATCH_FUNCTIONS      = {"open_excel", "close_excel"}


# ---------------------------------------------------------------------------
# Running Order generation from template read result
# ---------------------------------------------------------------------------

def generate_from_template(
    template_result,          # TemplateReadResult from the Template Reader module
    manifest: dict,           # filename -> {label, shape_type, ...} from cache
) -> list[dict]:
    """
    Build a list of Running Order row dicts from a TemplateReadResult.

    For each placeholder, based on content_type:
      "chart"   → insert_chart row (chart_type_ref blank; user selects from dropdown)
      "picture" → insert_picture row (image_path populated from placeholder)
      "excel"   → insert_from_excel row (excel_path, export_range, driver_range populated)
      ""        → empty_placeholder row

    For each unique Excel workbook referenced, a paired open_excel (scope=batch_open)
    and close_excel (scope=batch_close) row is added — one pair per workbook.

    Returns the full list including create_ppt header and save_ppt/save_pdf footer.
    """
    rows = []
    row_id = 1

    def _blank_row(func, note="", scope="normal"):
        return {
            "row_id": row_id, "enabled": 1, "scope": scope, "function": func,
            "slide_index": "", "placeholder": "", "chart_type_ref": "",
            "cache_file": "", "populations": "",
            "image_path": "", "excel_path": "", "export_range": "", "driver_range": "",
            "left_emu": "", "top_emu": "", "width_emu": "", "height_emu": "",
            "notes": note,
        }

    # --- Collect unique Excel workbooks to generate batch open/close pairs ---
    excel_paths_seen = []   # ordered, deduped
    for ph in template_result.placeholders:
        if ph.content_type == "excel" and ph.excel_path:
            if ph.excel_path not in excel_paths_seen:
                excel_paths_seen.append(ph.excel_path)

    # batch_open rows — absolute top, before all per-report rows
    for ep in excel_paths_seen:
        rows.append({
            "row_id": row_id, "enabled": 1, "scope": "batch_open", "function": "open_excel",
            "slide_index": "", "placeholder": "", "chart_type_ref": "",
            "cache_file": "", "populations": "",
            "image_path": "", "excel_path": ep, "export_range": "", "driver_range": "",
            "left_emu": "", "top_emu": "", "width_emu": "", "height_emu": "",
            "notes": f"Open workbook for batch: {os.path.basename(ep)}",
        })
        row_id += 1

    # --- Header rows (per-report) ---
    rows.append(_blank_row("create_ppt", "Open template and save working copy"))
    row_id += 1

    rows.append({**_blank_row("set_default_populations",
                              "Default populations for all charts — override per row in the populations column"),
                 "populations": "All^Selected"})
    row_id += 1

    rows.append(_blank_row("update_text", "Replace flag tokens with reporting unit values"))
    row_id += 1

    # --- Build URL → cache filename lookup ---
    url_to_cache = _build_url_to_cache_map(manifest)

    # --- One row per placeholder ---
    for ph in sorted(template_result.placeholders, key=lambda p: (p.slide_index, p.name)):
        ct = ph.content_type

        base = {
            "row_id": row_id, "enabled": 1, "scope": "normal",
            "slide_index": ph.slide_index, "placeholder": ph.name,
            "chart_type_ref": "", "cache_file": "", "populations": "",
            "image_path": "", "excel_path": "", "export_range": "", "driver_range": "",
            "left_emu": ph.left, "top_emu": ph.top,
            "width_emu": ph.width, "height_emu": ph.height,
            "notes": "",
        }

        if ct == "chart":
            cache_file = url_to_cache.get(_normalise_url(ph.url), "")
            rows.append({**base,
                "function": "insert_chart",
                "cache_file": cache_file,
                "notes": ph.label or "",
            })

        elif ct == "picture":
            rows.append({**base,
                "function": "insert_picture",
                "image_path": ph.image_path,
                "notes": f"Picture: {os.path.basename(ph.image_path)}",
            })

        elif ct == "excel":
            rows.append({**base,
                "function": "insert_from_excel",
                "excel_path":   ph.excel_path,
                "export_range": ph.excel_export_range,
                "driver_range": ph.excel_driver_range,
                "notes": f"Excel export: {ph.excel_export_range} from {os.path.basename(ph.excel_path)}",
            })

        else:
            rows.append({**base, "function": "empty_placeholder"})

        row_id += 1

    # --- Footer rows (per-report) ---
    rows.append({
        "row_id": row_id, "enabled": 1, "scope": "normal", "function": "save_ppt",
        "slide_index": "", "placeholder": "", "chart_type_ref": "",
        "cache_file": "", "populations": "",
        "image_path": "", "excel_path": "", "export_range": "", "driver_range": "",
        "left_emu": "", "top_emu": "", "width_emu": "", "height_emu": "",
        "notes": "Save output as .pptx",
    })
    row_id += 1

    rows.append({
        "row_id": row_id, "enabled": 0, "scope": "normal", "function": "save_pdf",
        "slide_index": "", "placeholder": "", "chart_type_ref": "",
        "cache_file": "", "populations": "",
        "image_path": "", "excel_path": "", "export_range": "", "driver_range": "",
        "left_emu": "", "top_emu": "", "width_emu": "", "height_emu": "",
        "notes": "Save output as .pdf (requires PowerPoint installed)",
    })
    row_id += 1

    # batch_close rows — absolute bottom, after all per-report rows
    for ep in excel_paths_seen:
        rows.append({
            "row_id": row_id, "enabled": 1, "scope": "batch_close", "function": "close_excel",
            "slide_index": "", "placeholder": "", "chart_type_ref": "",
            "cache_file": "", "populations": "",
            "image_path": "", "excel_path": ep, "export_range": "", "driver_range": "",
            "left_emu": "", "top_emu": "", "width_emu": "", "height_emu": "",
            "notes": f"Close workbook after batch: {os.path.basename(ep)}",
        })
        row_id += 1

    return rows


def _normalise_url(url: str) -> str:
    """Strip trailing slashes and whitespace for comparison."""
    return url.strip().rstrip("/")


def _build_url_to_cache_map(manifest: dict) -> dict:
    """Build url -> cache_filename map from the manifest."""
    mapping = {}
    for filename, entry in manifest.items():
        url = entry.get("url", "")
        if url:
            mapping[_normalise_url(url)] = filename
    return mapping


# ---------------------------------------------------------------------------
# Write Running Order to .xlsx
# ---------------------------------------------------------------------------

def write_xlsx(rows: list[dict], output_path: str,
               chart_type_map_path: str = None,
               manifest: dict = None):
    """Write Running Order rows to a formatted .xlsx file with dropdown validation on function, chart_type_ref, and enabled."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to write the Running Order xlsx.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Running Order"

    # --- Styles ---
    NAVY = "071A34"
    CRIMSON = "C12958"
    LIGHT_GREY = "F2F2F2"
    MID_GREY = "D9D9D9"
    WHITE = "FFFFFF"
    DISABLED_GREY = "AAAAAA"
    CHART_GREEN = "E8F5E9"
    PICTURE_TEAL = "E0F7FA"
    EXCEL_PURPLE = "F3E5F5"
    BATCH_ORANGE = "FFF3E0"
    STRUCTURAL_BLUE = "E3F2FD"
    POPULATIONS_AMBER = "FFF8E1"

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color=WHITE, bold=True, size=10)
    structural_fill = PatternFill("solid", fgColor=STRUCTURAL_BLUE)
    chart_fill = PatternFill("solid", fgColor=CHART_GREEN)
    disabled_font = Font(color=DISABLED_GREY, size=10)
    normal_font = Font(size=10)
    centre_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color=MID_GREY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Column widths ---
    col_widths = {
        "row_id":        6,
        "enabled":       8,
        "scope":         13,
        "function":      22,
        "slide_index":   11,
        "placeholder":   18,
        "chart_type_ref":22,
        "cache_file":    30,
        "populations":   30,
        "image_path":    36,
        "excel_path":    36,
        "export_range":  18,
        "driver_range":  18,
        "left_emu":      12,
        "top_emu":       12,
        "width_emu":     12,
        "height_emu":    12,
        "notes":         40,
    }

    # --- Header row ---
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centre_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # --- Build chart type lists per shape_type ---
    chart_type_by_shape = {}  # shape_type -> list of chart_type_refs
    if chart_type_map_path and os.path.exists(chart_type_map_path):
        with open(chart_type_map_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                st = row.get("data_shape", "").strip()
                ref = row.get("chart_type_ref", "").strip()
                if st and ref:
                    chart_type_by_shape.setdefault(st, []).append(ref)

    # Build cache_file -> shape_type from manifest
    cache_to_shape = {}
    if manifest:
        for filename, entry in manifest.items():
            st = entry.get("shape_type", "")
            if st:
                cache_to_shape[filename] = st

    # --- Data validation: function dropdown (applies to entire function column) ---
    func_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(ALL_FUNCTIONS) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(func_dv)

    enabled_dv = DataValidation(
        type="list",
        formula1='"1,0"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(enabled_dv)

    scope_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(SCOPE_VALUES) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(scope_dv)

    # --- Data rows ---
    for data_row_idx, row in enumerate(rows, start=2):
        excel_row = data_row_idx
        is_enabled = str(row.get("enabled", "1")) == "1"
        func  = row.get("function", "")
        scope = str(row.get("scope", "normal")).strip()
        is_structural = func in STRUCTURAL_FUNCTIONS

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = row.get(col_name, "")
            if value == "" or value is None:
                value = ""
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = border
            cell.font = disabled_font if not is_enabled else normal_font

            if scope in ("batch_open", "batch_close"):
                cell.fill = PatternFill("solid", fgColor=BATCH_ORANGE)
            elif is_structural and func == "set_default_populations":
                cell.fill = PatternFill("solid", fgColor=POPULATIONS_AMBER)
            elif is_structural:
                cell.fill = structural_fill
            elif func == "insert_chart":
                cell.fill = chart_fill
            elif func == "insert_picture":
                cell.fill = PatternFill("solid", fgColor=PICTURE_TEAL)
            elif func in ("insert_from_excel",):
                cell.fill = PatternFill("solid", fgColor=EXCEL_PURPLE)
            else:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)

            # Alignment
            if col_name in ("row_id", "enabled", "slide_index",
                            "left_emu", "top_emu", "width_emu", "height_emu"):
                cell.alignment = centre_align
            else:
                cell.alignment = left_align

        # Apply function dropdown
        func_col = COLUMNS.index("function") + 1
        func_dv.add(ws.cell(row=excel_row, column=func_col))

        # Apply enabled dropdown
        enabled_col = COLUMNS.index("enabled") + 1
        enabled_dv.add(ws.cell(row=excel_row, column=enabled_col))

        # Apply scope dropdown
        scope_col = COLUMNS.index("scope") + 1
        scope_dv.add(ws.cell(row=excel_row, column=scope_col))

        # Per-row chart_type_ref dropdown
        # Constrained to valid chart refs for the row's shape type if known;
        # falls back to all chart refs if cache_file is blank or shape unknown.
        cache_file = str(row.get("cache_file") or "").strip()
        if cache_file.lower() == "none":
            cache_file = ""
        shape_type = cache_to_shape.get(cache_file, "")
        chart_refs = chart_type_by_shape.get(shape_type, [])

        if not chart_refs:
            # Fallback: all chart refs across all shape types
            chart_refs = [
                ref
                for refs in chart_type_by_shape.values()
                for ref in refs
            ]

        if chart_refs and func == "insert_chart":
            ref_formula = '"' + ",".join(chart_refs) + '"'
            chart_dv = DataValidation(
                type="list",
                formula1=ref_formula,
                allow_blank=True,
                showDropDown=False,
            )
            ws.add_data_validation(chart_dv)
            ctr_col = COLUMNS.index("chart_type_ref") + 1
            chart_dv.add(ws.cell(row=excel_row, column=ctr_col))

        ws.row_dimensions[excel_row].height = 18

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Read Running Order from .xlsx
# ---------------------------------------------------------------------------

def read_xlsx(path: str) -> list[dict]:
    """
    Read the Running Order .xlsx and return a list of row dicts.
    Skips the header row and any row where enabled == 0.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to read the Running Order xlsx.")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = []
    for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v is not None for v in row):
            continue  # skip empty rows
        row_dict = {col: (row[i] if i < len(row) else "") for i, col in enumerate(COLUMNS)}
        row_dict.setdefault("enabled", "1")
        coerce_row(row_dict)
        # Normalise scope — default to "normal" if blank or missing
        scope = str(row_dict.get("scope", "")).strip()
        if scope not in SCOPE_VALUES:
            scope = "normal"
        row_dict["scope"] = scope
        rows.append(row_dict)

    return rows


# ---------------------------------------------------------------------------
# Convenience: write CSV fallback (no openpyxl)
# ---------------------------------------------------------------------------

def write_csv(rows: list[dict], output_path: str):
    """Write Running Order rows to a plain CSV (no formatting)."""
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
